import mysql.connector
from mysql.connector import RefreshOption
import openpyxl
import os


class Segment:
    def __init__(self, segment_ID, time, speed, angle, exercise_ID,direction):
        self.segment_ID = segment_ID
        self.time = time
        self.speed = speed
        self.angle = angle
        self.exercise_ID = exercise_ID
        self.direction = direction

class exercise:
    def __init__(self, exercise_ID, exercise_name, description, time):
        self.exercise_ID = exercise_ID
        self.exercise_name = exercise_name
        self.description = description
        self.time = time

class user:
    def __init__(self, user_id, gender, age, name,):
        self.user_id = user_id
        self.gender = gender
        self.age = age
        self.name = name
         



def connect_myc():
    db = mysql.connector.connect(
     host="localhost",
    user="root",
    passwd="1995",
    database="dfg"
    )
    refresh = RefreshOption.LOG
    db.cmd_refresh(refresh)
    myc = db.cursor()
    return db, myc


def get_exercise_progrems(myc, exercise_ID): # get the exercise progrems from the database

    myc.execute( f"SELECT * FROM exercise_sement WHERE exercise_ID = {exercise_ID} ORDER BY segment_time")    
    myresult = myc.fetchall()
    Segment_list = []
    for row in myresult:
        s = Segment(row[0], row[1], row[2], row[3], row[4], row[5])
        Segment_list.append(s)
        #print("ID", segment.segment_ID,"time", segment.time,"speed", segment.speed,"angle", segment.angle,"exercise_ID",  segment.exercise_ID,"direction",  segment.direction)
    myc.execute( f"SELECT * FROM exercise_progrems WHERE exercise_ID = {exercise_ID}")
    myresult = myc.fetchall()

    exercise_progrem = exercise(myresult[0][0], myresult[0][1], myresult[0][2], myresult[0][3])
    print("ID", exercise_progrem.exercise_ID,"name", exercise_progrem.exercise_name,"description", exercise_progrem.description,"time",  exercise_progrem.time)

    return Segment_list, exercise_progrem

def get_user(myc, user_id): # get the user from the database
    myc.execute( f"SELECT * FROM users WHERE user_id = {user_id}")
    myresult = myc.fetchall()
    user1 = user(myresult[0][0], myresult[0][1], myresult[0][2], myresult[0][3])
    print("ID", user1.user_id,"name", user1.name)
    return user1
    




def calibrate_body_angle(sheet):
    # take all the angel of the body from the exel file and calculate the average
    shoulder = []
    torso_RL = []
    torso_BF = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4, values_only=True):
        shoulder.append(row[0])
        torso_RL.append(row[1])
        torso_BF.append(row[2])
    # calculate the average of the angel
    shoulder_avg = sum(shoulder) / len(shoulder)
    torso_RL_avg = sum(torso_RL) / len(torso_RL)
    torso_BF_avg = sum(torso_BF) / len(torso_BF)
    return shoulder_avg, torso_RL_avg, torso_BF_avg


def create_chart(sheet):
        #create the chart in the excel file time will be the x axis and the angel will be the y axis
    chart = openpyxl.chart.LineChart()
    chart.style = 10
    chart.y_axis.title = "Angel"
    chart.x_axis.title = "Time"
    chart.title = "Angel over time"
    data = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row, max_col=sheet.max_column)
    chart.add_data(data, titles_from_data=True)
    sheet.add_chart(chart, "F1")


def create_user_directory(user_name):
    directory_path = os.path.join(os.getcwd(), user_name)
    try:
        os.makedirs(directory_path, exist_ok=True)
        print(f"Directory '{directory_path}' created successfully.")
    except Exception as e:
        print(f"Failed to create directory '{directory_path}'. Error: {e}")
    return directory_path

def create_plot():
    # create the plot for the angel
    import matplotlib.pyplot as plt
    plt.ion()
    plt.show()
    plt.title('Angel over time')
    plt.xlabel('Time')
    plt.ylabel('Angel')
    plt.grid()
    plt.plot([], [], 'ro', label='shoulder')
    plt.plot([], [], 'bo', label='torso_RL')
    plt.plot([], [], 'go', label='torso_BF')

    plt.ylim(-20, 20)

    plt.legend()

    return plt

def save_plot(plt, shoulder, torso_RL, torso_BF):
    # save the plot to the directory
    plt.plot(shoulder, label='shoulder')
    plt.plot(torso_RL, label='torso_RL')
    plt.plot(torso_BF, label='torso_BF')
    plt.xlabel('Time')
    plt.ylabel('Angel')
    plt.legend()
    plt.savefig('plot.png')
    plt.pause(0.05)
    plt.clf()

def update_plot(plt, timer, shoulder, torso_RL, torso_BF, platform_angle_BF, platform_angle_RL, shoulder_avg, torso_RL_avg, torso_BF_avg):
    # update the plot with the new data
    plt.plot(timer, shoulder, 'ro')
    plt.plot(timer, torso_RL, 'bo')
    plt.plot(timer, torso_BF, 'go')
    plt.pause(0.0000005)

    # save only the last 100 points
    if timer > 3:
        plt.xlim(timer - 3, timer)
        # delete the first point
        plt.gca().lines[0].remove()
        plt.gca().lines[1].remove()
        plt.gca().lines[2].remove()


    else:
        plt.xlim(0, 3)

def print_plot(plt, angel, angel_avg, platform_angel, direction):
    # take all the angel of the body from the exel file of the last 60 sec and print the plot
    plt.clf()
    plt.ylim(-20, 20)
    plt.grid()

    if direction == 'f' or direction == 'b':
        # תיצור גרף שמציג את הנקודות ברגע נתון אחד בלי ציר זמן ואז תוסיף את הקווים של הממוצעים
        plt.plot(angel[2], label='torso_B\F', color='r', marker='o')
        plt.plot(angel_avg[0], label='torso_B\F_avg', color='b', marker='x')
        plt.plot(platform_angel[0], label='platform_angle_B\F', color='g', marker='v')

    if direction == 'r' or direction == 'l':
        plt.plot(angel[1], label='torso_R\L', color='r', marker='o')
        plt.plot(angel[0], label='shoulder', color='b', marker='o')
        plt.plot(angel_avg[0], label='shoulder_avg', color='b', marker='x')
        plt.plot(angel_avg[1], label='torso_R\L_avg', color='r', marker='x')
        plt.plot(platform_angel[1], label='platform_angle_R\L', color='g', marker='o')

    plt.xlabel('Time')
    plt.ylabel('Angel')

    plt.legend()
    # plt.pause(10)



if __name__ == "__main__":
    plt = create_plot()
    angel = [1, 2, -1]
    angel_avg = [1.5, 1.7, -0.5]
    platform_angle = [5, 0]
    print_plot(plt, angel, angel_avg, platform_angle, 'f')
